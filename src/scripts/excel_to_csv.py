#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook

# 定义输入和输出路径
excel_file = '/Users/zhangzekun/code/anaysis/ll2/data/异常用电&线损综合分析数据模型V1.0.xlsx'
output_dir = '/Users/zhangzekun/code/anaysis/ll2/csv_output'

# 确保输出目录存在
os.makedirs(output_dir, exist_ok=True)

# 加载工作簿以获取所有工作表名称
print(f"正在加载Excel文件: {excel_file}")
workbook = load_workbook(excel_file, read_only=True, data_only=True)
sheet_names = workbook.sheetnames
print(f"找到 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")

# 关闭工作簿
workbook.close()

# 处理每个工作表
for sheet_name in sheet_names:
    print(f"\n处理工作表: {sheet_name}")
    
    # 尝试读取工作表
    try:
        # 首先尝试直接读取整个工作表
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # 检查是否为空工作表
        if df.empty:
            print(f"  工作表 '{sheet_name}' 为空，跳过")
            continue
        
        # 检查是否有多个表格（通过查找连续的空行来判断）
        # 获取全部为NaN的行的索引
        empty_rows = df.index[df.isnull().all(axis=1)].tolist()
        
        # 如果没有空行或只有少量空行，认为是单个表格
        if len(empty_rows) <= 2:
            # 保存为单个CSV文件
            csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
            df.to_csv(csv_file, index=False, encoding='utf-8-sig')
            print(f"  已保存单个表格到: {csv_file}")
        else:
            # 可能有多个表格，查找连续的非空行块
            table_start_indices = [0]  # 第一个表格从索引0开始
            
            # 查找可能的表格分隔点（连续的空行后的第一个非空行）
            for i in range(1, len(empty_rows)):
                if empty_rows[i] - empty_rows[i-1] > 1:  # 不连续的空行
                    table_start_indices.append(empty_rows[i-1] + 1)
            
            # 添加最后一个表格的开始索引（如果有）
            if empty_rows and empty_rows[-1] + 1 < len(df):
                table_start_indices.append(empty_rows[-1] + 1)
            
            # 如果只找到一个起始点，可能是单个表格带有空行
            if len(table_start_indices) <= 1:
                # 保存为单个CSV文件，但去除全空行
                df_clean = df.dropna(how='all')
                csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
                df_clean.to_csv(csv_file, index=False, encoding='utf-8-sig')
                print(f"  已保存清理后的单个表格到: {csv_file}")
            else:
                # 处理多个表格
                print(f"  检测到可能的多个表格，分别保存")
                
                for table_idx, start_idx in enumerate(table_start_indices):
                    # 确定表格结束索引
                    if table_idx < len(table_start_indices) - 1:
                        end_idx = table_start_indices[table_idx + 1] - 2  # 减去空行
                    else:
                        end_idx = len(df) - 1
                    
                    # 提取表格数据
                    table_df = df.iloc[start_idx:end_idx+1].copy()
                    
                    # 跳过空表格
                    if table_df.dropna(how='all').empty:
                        continue
                    
                    # 清理表格（删除全空行和全空列）
                    table_df = table_df.dropna(how='all')
                    table_df = table_df.dropna(axis=1, how='all')
                    
                    # 如果表格仍然为空，跳过
                    if table_df.empty:
                        continue
                    
                    # 保存为CSV
                    csv_file = os.path.join(output_dir, f"{sheet_name}_表{table_idx+1}.csv")
                    table_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                    print(f"  已保存表格 {table_idx+1} 到: {csv_file}")
    
    except Exception as e:
        print(f"  处理工作表 '{sheet_name}' 时出错: {str(e)}")

print("\n所有工作表处理完成！CSV文件已保存到目录: {}".format(output_dir))